#percorrer a base de dados
#p/ cada item
    #ver se o bairro já existe em uma aba, senão, criar uma
    #copiar a linha e add os valores a ela

import openpyxl as oxl
#import do modulo para copiar infos de objetos em python e cola-los corretamente
from copy import copy


def criar_aba(bairro, bairros_file, aba_baseDados, estilo_cabecalho):
    #se nao houver uma aba com um determinado bairro, crie-o
    if bairro not in bairros_file.sheetnames:
        bairros_file.create_sheet(bairro)
        nova_aba = bairros_file[bairro]

        #IGUALANDO OS VALORES DA LINHA 1 (CATEGORIAS) NO EXCEL
        ultima_coluna = len(aba_baseDados["1"])
        for coluna in range(1, ultima_coluna+1):
            nova_aba.cell(row=1, column=coluna).value = aba_baseDados.cell(row=1, column=coluna).value
            nova_aba.cell(row=1, column=coluna)._style = estilo_cabecalho



def transferir_dados_aba(aba_origem, aba_destino, linha_origem):
    #indicando que o conteudo a ser preenchido sera 1 posicao depois da ultima linha preenchida
    linha_destino = aba_destino.max_row + 1
    for coluna in range(1, 4): #pra cada coluna (1-3)
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna) #diz qual a posicao e valor da celula de origem, de onde tera como referencia o valor
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna) #diz qual a posicao e valor da celula de destino, aonde sera atribuido o valor
        celula_destino.value = celula_origem.value #atribui o valor
        celula_destino._style = copy(celula_origem._style)


#start file
bairros_file = oxl.load_workbook('Bairros.xlsx')

aba_baseDados = bairros_file["Base de Dados"]
print(aba_baseDados)

#forma facil de requisitar o valor da ultima linha pra se saber a qtd de dados
#mas nao muito confiavel, entao ficar de olho
ultima_linha = len(aba_baseDados["A"])
print(ultima_linha)
estilo_cabecalho = copy(aba_baseDados["A1"]._style) #formatacao da primeira celula

for linha in range(2, ultima_linha+1):
    #a partir da aba de dados geral, captura o bairro, que esta na coluna C 
    bairro = aba_baseDados[f"C{linha}"].value
    #outra forma, pegando na celula onde a linha = linha do ciclo e a coluna 3
    bairro = aba_baseDados.cell(row=linha, column=3).value
    
    #caso de celula vazia
    if not bairro:
        break

    #cria uma aba p/ bairro
    criar_aba(bairro, bairros_file, aba_baseDados, estilo_cabecalho)

    #apos a criaçao da aba pelo metodo acima, eh atribuido seu valor para a var abaixo
    aba_destino = bairros_file[bairro]

    #p/ ocorrer a transferencia, necessita-se mandar de onde os dados vem, para onde vao e qual a linha(valores) que serao enviados
    transferir_dados_aba(aba_baseDados, aba_destino, linha)

bairros_file.save("Bairros2.xlsx")
